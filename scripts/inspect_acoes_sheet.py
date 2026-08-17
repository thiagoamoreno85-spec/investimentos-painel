from pathlib import Path
import json
from openpyxl import load_workbook

WORKBOOK = Path('/home/ubuntu/projects/investimentos-ac212e00/CONTROLE INVESTIMENTOS - TAM .xlsx')
OUTPUT = Path('/home/ubuntu/investimentos-painel/.tmp_acoes_sheet.json')

wb = load_workbook(WORKBOOK, data_only=False, read_only=True)
ws = wb['AÇÕES ']

rows = []
for row in ws.iter_rows():
    values = []
    for cell in row:
        value = cell.value
        if value is not None:
            values.append({
                'coordinate': cell.coordinate,
                'value': value,
                'number_format': cell.number_format,
            })
    if values:
        rows.append(values)

payload = {
    'title': ws.title,
    'max_row': ws.max_row,
    'max_column': ws.max_column,
    'rows': rows,
}
OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, default=str, indent=2))
print(f'Arquivo salvo em {OUTPUT}')
print(f'Aba: {ws.title}; linhas: {ws.max_row}; colunas: {ws.max_column}; linhas não vazias: {len(rows)}')

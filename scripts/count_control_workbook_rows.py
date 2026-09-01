from pathlib import Path
import json
import openpyxl

workbook_path = Path('/home/ubuntu/upload/CONTROLEINVESTIMENTOS-TAM.xlsx')
workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)

sheet_summary = []
total_rows = 0
for name in workbook.sheetnames:
    worksheet = workbook[name]
    row_count = sum(1 for _ in worksheet.iter_rows(min_row=2, values_only=True))
    total_rows += row_count
    sheet_summary.append({'sheet': name, 'rows_after_header': row_count})

workbook.close()
print(json.dumps({'path': str(workbook_path), 'total_rows_after_headers': total_rows, 'sheets': sheet_summary}, ensure_ascii=False, indent=2))

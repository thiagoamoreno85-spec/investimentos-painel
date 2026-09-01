from pathlib import Path
from datetime import datetime, date
import json
import openpyxl

WORKBOOK_PATH = Path('/home/ubuntu/upload/CONTROLEINVESTIMENTOS-TAM.xlsx')
OUTPUT_PATH = Path('/home/ubuntu/investimentos-painel/.tmp_sep01_control_records.json')
KEYWORDS = ('TWST', 'URNM', 'INDA', 'BTC', 'ETH', 'USDT', 'SELIC', 'TESOURO', 'CDB')
TARGET_DATE = date(2026, 9, 1)

def display(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value

workbook = openpyxl.load_workbook(WORKBOOK_PATH, read_only=True, data_only=False)
matches = []

for sheet_name in workbook.sheetnames:
    worksheet = workbook[sheet_name]
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        normalized = [display(value) for value in row]
        row_text = ' | '.join(str(value).upper() for value in normalized if value is not None)
        has_keyword = any(keyword in row_text for keyword in KEYWORDS)
        has_target_date = any(
            isinstance(value, (datetime, date)) and value.date() == TARGET_DATE if isinstance(value, datetime) else isinstance(value, date) and value == TARGET_DATE
            for value in row
        )
        if has_keyword or has_target_date:
            matches.append({
                'sheet': sheet_name,
                'row': row_number,
                'values': normalized,
            })

workbook.close()
OUTPUT_PATH.write_text(json.dumps(matches, ensure_ascii=False, indent=2, default=str), encoding='utf-8')
print(f'Matches: {len(matches)}')
print(f'Output: {OUTPUT_PATH}')

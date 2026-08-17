from pathlib import Path
import csv
from openpyxl import load_workbook

WORKBOOK = Path('/home/ubuntu/projects/investimentos-ac212e00/CONTROLE INVESTIMENTOS - TAM .xlsx')
OUTPUT = Path('/home/ubuntu/investimentos-painel/.tmp_acoes_positions.csv')

wb = load_workbook(WORKBOOK, data_only=True, read_only=True)
ws = wb['AÇÕES ']

rows = []
for row_number in range(15, 117):
    ticker = ws[f'G{row_number}'].value
    quantity = ws[f'H{row_number}'].value
    cost = ws[f'J{row_number}'].value
    price = ws[f'L{row_number}'].value
    if isinstance(ticker, str) and ticker.strip() and quantity not in (None, 0):
        rows.append({
            'row': row_number,
            'ticker': ticker.strip(),
            'quantity': quantity,
            'average_cost': cost,
            'price_reference': price,
        })

with OUTPUT.open('w', newline='', encoding='utf-8') as handle:
    writer = csv.DictWriter(handle, fieldnames=['row', 'ticker', 'quantity', 'average_cost', 'price_reference'])
    writer.writeheader()
    writer.writerows(rows)

print(f'Posições extraídas: {len(rows)}')
for position in rows:
    print('{row:>3} | {ticker:<24} | qtd={quantity} | custo={average_cost} | preço={price_reference}'.format(**position))

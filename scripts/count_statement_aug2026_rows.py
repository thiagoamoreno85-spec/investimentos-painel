from pathlib import Path
import json
import openpyxl

statement_path = Path('/home/ubuntu/upload/Extrato496056AGO.2026.xlsx')
workbook = openpyxl.load_workbook(statement_path, read_only=True, data_only=True)

sheets = []
total_rows = 0
for name in workbook.sheetnames:
    worksheet = workbook[name]
    rows_after_header = sum(1 for _ in worksheet.iter_rows(min_row=2, values_only=True))
    total_rows += rows_after_header
    sheets.append({'sheet': name, 'rows_after_header': rows_after_header})

workbook.close()
print(json.dumps({'total_rows_after_headers': total_rows, 'sheets': sheets}, ensure_ascii=False, indent=2))
